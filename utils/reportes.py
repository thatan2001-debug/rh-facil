"""
Reportes y dashboard avanzados de Gestor RH IA (Etapa Opción D).

Aprovecha los datos de:
- Ficha ampliada del empleado (Etapa 5)
- Historial de documentos
- Múltiples motivos de retiro

Métricas incluidas:
- Distribución por tipo de contrato
- Distribución por modalidad (presencial/remoto/híbrido)
- Distribución por área/sede
- Rotación mensual (ingresos vs retiros)
- Antigüedad promedio
- Costos laborales estimados con carga prestacional
- Documentos generados por tipo y por mes
- Empleados con contratos por vencer
"""

import streamlit as st
import pandas as pd
from datetime import datetime, timedelta, date
from collections import Counter, defaultdict


def _parse_fecha(fecha_str) -> datetime:
    """Parsea fecha en múltiples formatos. Retorna None si falla."""
    if not fecha_str:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(str(fecha_str)[:10], fmt)
        except ValueError:
            continue
    return None


def _dias_entre(d1, d2) -> int:
    """Días entre dos fechas (int)."""
    if not d1 or not d2:
        return 0
    return abs((d2 - d1).days)


# ══════════════════════════════════════════════════════════════════════════════
# CÁLCULO DE MÉTRICAS
# ══════════════════════════════════════════════════════════════════════════════

def calcular_metricas_empleados(empleados: list) -> dict:
    """Genera todas las métricas a partir de la lista de empleados."""
    if not empleados:
        return {"total": 0, "activos": 0, "retirados": 0}

    activos = [e for e in empleados if e.get("activo", True)]
    retirados = [e for e in empleados if not e.get("activo", True)]

    # ── Distribución por tipo de contrato ─────────────────────────
    tipos_contrato = Counter()
    for e in activos:
        tipos_contrato[str(e.get("tipo_contrato", "Sin definir")).capitalize()] += 1

    # ── Modalidad ─────────────────────────────────────────────────
    modalidades = Counter()
    for e in activos:
        mod = str(e.get("modalidad", "presencial")).capitalize()
        modalidades[mod] += 1

    # ── Áreas ─────────────────────────────────────────────────────
    areas = Counter()
    for e in activos:
        area = str(e.get("area", "") or "Sin área").strip() or "Sin área"
        areas[area] += 1

    # ── Sedes ─────────────────────────────────────────────────────
    sedes = Counter()
    for e in activos:
        sede = str(e.get("sede", "") or "Sin sede").strip() or "Sin sede"
        sedes[sede] += 1

    # ── Antigüedad ────────────────────────────────────────────────
    hoy = datetime.today()
    dias_totales = []
    for e in activos:
        fi = _parse_fecha(e.get("fecha_ingreso"))
        if fi:
            dias_totales.append(_dias_entre(fi, hoy))
    antiguedad_promedio_dias = int(sum(dias_totales) / len(dias_totales)) if dias_totales else 0
    antiguedad_promedio_años = antiguedad_promedio_dias / 365.25

    # ── Rangos de antigüedad ──────────────────────────────────────
    rangos_antiguedad = Counter()
    for d in dias_totales:
        años = d / 365.25
        if años < 1:
            rangos_antiguedad["Menos de 1 año"] += 1
        elif años < 3:
            rangos_antiguedad["1-3 años"] += 1
        elif años < 5:
            rangos_antiguedad["3-5 años"] += 1
        elif años < 10:
            rangos_antiguedad["5-10 años"] += 1
        else:
            rangos_antiguedad["Más de 10 años"] += 1

    # ── Costos laborales ──────────────────────────────────────────
    total_nomina = sum(float(e.get("salario", 0) or 0) for e in activos)
    total_variable = sum(float(e.get("ingreso_promedio_variable", 0) or 0) for e in activos)
    total_aux_transp = sum(float(e.get("auxilio_transporte", 0) or 0) for e in activos)

    # Prestaciones: aproximadamente 21.35% del salario
    # (cesantías 8.33% + intereses 1% + prima 8.33% + vacaciones 4.17% = 21.83%)
    prestacional = total_nomina * 0.2135

    # Aportes patronales: aprox 20.5% (salud 8.5% + pensión 12%)
    aportes = total_nomina * 0.205

    total_carga = total_nomina + total_variable + total_aux_transp + prestacional + aportes

    # ── Contratos por vencer (próximos 30 días) ────────────────
    contratos_por_vencer = []
    limite_30d = hoy + timedelta(days=30)
    for e in activos:
        fv = _parse_fecha(e.get("fecha_vencimiento_contrato"))
        if fv and hoy <= fv <= limite_30d:
            dias_restantes = _dias_entre(hoy, fv)
            contratos_por_vencer.append({
                "nombre": e.get("nombre", ""),
                "documento": e.get("documento", ""),
                "tipo_contrato": e.get("tipo_contrato", ""),
                "dias_restantes": dias_restantes,
                "fecha_vencimiento": e.get("fecha_vencimiento_contrato", ""),
            })
    contratos_por_vencer.sort(key=lambda x: x["dias_restantes"])

    # ── Rotación: ingresos y retiros por mes (últimos 12) ──────
    rotacion_mensual = defaultdict(lambda: {"ingresos": 0, "retiros": 0})
    hoy_mes = hoy.replace(day=1)

    # Generar los últimos 12 meses
    for i in range(12):
        mes = (hoy_mes.replace(day=1) - timedelta(days=30 * i)).strftime("%Y-%m")
        rotacion_mensual[mes]  # Inicializar

    for e in empleados:
        fi = _parse_fecha(e.get("fecha_ingreso"))
        if fi and fi >= hoy - timedelta(days=365):
            mes_key = fi.strftime("%Y-%m")
            if mes_key in rotacion_mensual:
                rotacion_mensual[mes_key]["ingresos"] += 1

        fr = _parse_fecha(e.get("fecha_retiro"))
        if fr and fr >= hoy - timedelta(days=365):
            mes_key = fr.strftime("%Y-%m")
            if mes_key in rotacion_mensual:
                rotacion_mensual[mes_key]["retiros"] += 1

    return {
        "total":                  len(empleados),
        "activos":                len(activos),
        "retirados":              len(retirados),
        "tipos_contrato":         dict(tipos_contrato),
        "modalidades":            dict(modalidades),
        "areas":                  dict(areas),
        "sedes":                  dict(sedes),
        "antiguedad_prom_dias":   antiguedad_promedio_dias,
        "antiguedad_prom_años":   antiguedad_promedio_años,
        "rangos_antiguedad":      dict(rangos_antiguedad),
        "nomina_fija":            total_nomina,
        "nomina_variable":        total_variable,
        "aux_transporte":         total_aux_transp,
        "prestacional":           prestacional,
        "aportes_patronales":     aportes,
        "costo_total":            total_carga,
        "contratos_por_vencer":   contratos_por_vencer,
        "rotacion_mensual":       dict(rotacion_mensual),
    }


def calcular_metricas_documentos(historial: list) -> dict:
    """Métricas basadas en el historial de documentos generados."""
    if not historial:
        return {"total": 0}

    # Por tipo
    por_tipo = Counter(h.get("tipo_documento", "otro") for h in historial)

    # Por mes (últimos 12)
    por_mes = defaultdict(int)
    hoy = datetime.today()
    for h in historial:
        fecha_str = str(h.get("generado_en") or h.get("fecha", ""))[:7]
        if fecha_str and fecha_str >= (hoy - timedelta(days=365)).strftime("%Y-%m"):
            por_mes[fecha_str] += 1

    # Enviados por correo
    enviados = sum(1 for h in historial if h.get("enviado_por_correo"))

    return {
        "total":     len(historial),
        "por_tipo":  dict(por_tipo),
        "por_mes":   dict(por_mes),
        "enviados":  enviados,
    }


# ══════════════════════════════════════════════════════════════════════════════
# UI: RENDERIZAR REPORTES
# ══════════════════════════════════════════════════════════════════════════════

def mostrar_reportes(email_empresa: str):
    """Renderiza la pantalla completa de reportes."""
    from utils.empleados_db import empleados_listar
    from utils.historial import obtener

    st.markdown("# 📊 Reportes y Análisis")
    st.caption("Métricas operativas y análisis de tu equipo humano")

    # Cargar datos
    empleados = empleados_listar(email_empresa)
    historial = obtener(email_empresa, limite=2000)

    if not empleados:
        st.info("📭 Aún no tienes empleados registrados. Ve a **👥 Empleados** para empezar.")
        return

    m_emp = calcular_metricas_empleados(empleados)
    m_doc = calcular_metricas_documentos(historial)

    # ── PANORAMA GENERAL ────────────────────────────────────────
    st.markdown("### 📈 Panorama general")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("👥 Total empleados", m_emp["total"])
    c2.metric("✅ Activos", m_emp["activos"])
    c3.metric("⭕ Retirados", m_emp["retirados"])
    c4.metric("📄 Documentos generados", m_doc.get("total", 0))

    st.divider()

    # ── COSTOS LABORALES ───────────────────────────────────────
    st.markdown("### 💵 Costos laborales estimados")
    st.caption("Cálculo mensual aproximado. Incluye salario, variables, auxilio de transporte, "
                 "prestaciones sociales (~21.35%) y aportes patronales (~20.5%).")

    cc1, cc2, cc3, cc4 = st.columns(4)
    cc1.metric("💰 Nómina fija",
                f"${m_emp['nomina_fija']:,.0f}".replace(",", "."))
    cc2.metric("📊 Variables",
                f"${m_emp['nomina_variable']:,.0f}".replace(",", "."))
    cc3.metric("🚌 Auxilio transporte",
                f"${m_emp['aux_transporte']:,.0f}".replace(",", "."))
    cc4.metric("📋 Prestaciones",
                f"${m_emp['prestacional']:,.0f}".replace(",", "."))

    st.info(
        f"💼 **Costo laboral mensual total estimado:** "
        f"${m_emp['costo_total']:,.0f} COP".replace(",", ".") + "  \n"
        f"Anual: ${m_emp['costo_total'] * 12:,.0f} COP".replace(",", ".")
    )

    st.divider()

    # ── CONTRATOS POR VENCER ────────────────────────────────────
    if m_emp["contratos_por_vencer"]:
        st.markdown(f"### ⚠️ Contratos por vencer en los próximos 30 días "
                      f"({len(m_emp['contratos_por_vencer'])})")
        df_venc = pd.DataFrame(m_emp["contratos_por_vencer"])
        df_venc = df_venc.rename(columns={
            "nombre":            "Empleado",
            "documento":         "Documento",
            "tipo_contrato":     "Tipo",
            "dias_restantes":    "Días restantes",
            "fecha_vencimiento": "Vencimiento",
        })
        st.dataframe(df_venc, use_container_width=True, hide_index=True)
        st.divider()

    # ── ANÁLISIS DE COMPOSICIÓN ─────────────────────────────────
    st.markdown("### 📊 Composición del equipo")

    c1, c2 = st.columns(2)

    with c1:
        st.markdown("**Por tipo de contrato**")
        if m_emp["tipos_contrato"]:
            df_tc = pd.DataFrame(
                list(m_emp["tipos_contrato"].items()),
                columns=["Tipo", "Cantidad"]
            )
            st.bar_chart(df_tc.set_index("Tipo"), height=300)
        else:
            st.info("Sin datos de tipo de contrato")

    with c2:
        st.markdown("**Por modalidad**")
        if m_emp["modalidades"]:
            df_mod = pd.DataFrame(
                list(m_emp["modalidades"].items()),
                columns=["Modalidad", "Cantidad"]
            )
            st.bar_chart(df_mod.set_index("Modalidad"), height=300)
        else:
            st.info("Sin datos de modalidad")

    c3, c4 = st.columns(2)

    with c3:
        st.markdown("**Por área**")
        if m_emp["areas"] and len(m_emp["areas"]) > 0:
            areas_top = dict(Counter(m_emp["areas"]).most_common(8))
            df_ar = pd.DataFrame(
                list(areas_top.items()),
                columns=["Área", "Cantidad"]
            )
            st.bar_chart(df_ar.set_index("Área"), height=300)
        else:
            st.info("Sin datos de área. Llena este campo en la ficha de empleados.")

    with c4:
        st.markdown("**Por sede**")
        if m_emp["sedes"] and len(m_emp["sedes"]) > 0:
            df_sd = pd.DataFrame(
                list(m_emp["sedes"].items()),
                columns=["Sede", "Cantidad"]
            )
            st.bar_chart(df_sd.set_index("Sede"), height=300)
        else:
            st.info("Sin datos de sede.")

    st.divider()

    # ── ANTIGÜEDAD ──────────────────────────────────────────────
    st.markdown("### 📅 Antigüedad del equipo")
    ca1, ca2 = st.columns([1, 2])

    with ca1:
        años = m_emp["antiguedad_prom_años"]
        meses_extra = int((años - int(años)) * 12)
        años_int = int(años)
        st.metric("Antigüedad promedio",
                    f"{años_int} año(s) {meses_extra} mes(es)")
        st.metric("Total días acumulados",
                    f"{m_emp['antiguedad_prom_dias']:,}".replace(",", "."))

    with ca2:
        st.markdown("**Distribución por rangos de antigüedad**")
        if m_emp["rangos_antiguedad"]:
            df_rango = pd.DataFrame(
                list(m_emp["rangos_antiguedad"].items()),
                columns=["Rango", "Empleados"]
            )
            st.bar_chart(df_rango.set_index("Rango"), height=250)

    st.divider()

    # ── ROTACIÓN ────────────────────────────────────────────────
    st.markdown("### 🔄 Rotación (últimos 12 meses)")
    if m_emp["rotacion_mensual"]:
        # Convertir a DataFrame
        meses_ordenados = sorted(m_emp["rotacion_mensual"].keys())
        df_rot = pd.DataFrame([
            {
                "Mes": mes,
                "Ingresos": m_emp["rotacion_mensual"][mes]["ingresos"],
                "Retiros": m_emp["rotacion_mensual"][mes]["retiros"],
            }
            for mes in meses_ordenados
        ])
        st.bar_chart(df_rot.set_index("Mes"), height=300)

        # Totales
        total_ing = df_rot["Ingresos"].sum()
        total_ret = df_rot["Retiros"].sum()
        neto = total_ing - total_ret
        cr1, cr2, cr3 = st.columns(3)
        cr1.metric("📈 Ingresos (12m)", total_ing)
        cr2.metric("📉 Retiros (12m)", total_ret)
        cr3.metric("📊 Neto", f"{'+' if neto >= 0 else ''}{neto}")
    else:
        st.info("Sin datos de rotación aún.")

    st.divider()

    # ── DOCUMENTOS GENERADOS ───────────────────────────────────
    if m_doc.get("total", 0) > 0:
        st.markdown("### 📄 Actividad documental")

        cd1, cd2 = st.columns(2)

        with cd1:
            st.markdown("**Documentos por tipo (Top 10)**")
            por_tipo = m_doc.get("por_tipo", {})
            if por_tipo:
                top = dict(Counter(por_tipo).most_common(10))
                df_dt = pd.DataFrame(
                    list(top.items()),
                    columns=["Tipo", "Cantidad"]
                )
                df_dt["Tipo"] = df_dt["Tipo"].str.replace("_", " ").str.title()
                st.bar_chart(df_dt.set_index("Tipo"), height=300)

        with cd2:
            st.markdown("**Documentos generados por mes**")
            por_mes = m_doc.get("por_mes", {})
            if por_mes:
                meses = sorted(por_mes.keys())
                df_dm = pd.DataFrame([
                    {"Mes": m, "Documentos": por_mes[m]}
                    for m in meses
                ])
                st.bar_chart(df_dm.set_index("Mes"), height=300)

        st.info(
            f"📊 **{m_doc.get('total', 0)}** documentos generados en total  \n"
            f"✉️ **{m_doc.get('enviados', 0)}** enviados por correo automáticamente"
        )

    # ── EXPORTAR ─────────────────────────────────────────────────
    st.divider()
    st.markdown("### 📥 Exportar reporte")

    if st.button("📊 Descargar métricas en Excel", type="primary"):
        _generar_excel_reporte(m_emp, m_doc)


def _generar_excel_reporte(m_emp: dict, m_doc: dict):
    """Genera un Excel con todas las métricas para descargar."""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # Hoja 1: Resumen
    ws = wb.active
    ws.title = "Resumen"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1B3F6E", end_color="1B3F6E", fill_type="solid")

    ws["A1"] = "Reporte de Gestor RH IA"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    fila = 4
    secciones = [
        ("PANORAMA GENERAL", [
            ("Total empleados", m_emp["total"]),
            ("Activos", m_emp["activos"]),
            ("Retirados", m_emp["retirados"]),
        ]),
        ("COSTOS LABORALES", [
            ("Nómina fija", f"${m_emp['nomina_fija']:,.0f}"),
            ("Variables", f"${m_emp['nomina_variable']:,.0f}"),
            ("Aux. transporte", f"${m_emp['aux_transporte']:,.0f}"),
            ("Prestaciones (21.35%)", f"${m_emp['prestacional']:,.0f}"),
            ("Aportes patronales (20.5%)", f"${m_emp['aportes_patronales']:,.0f}"),
            ("TOTAL MENSUAL", f"${m_emp['costo_total']:,.0f}"),
            ("TOTAL ANUAL", f"${m_emp['costo_total'] * 12:,.0f}"),
        ]),
        ("ANTIGÜEDAD", [
            ("Promedio (años)", f"{m_emp['antiguedad_prom_años']:.1f}"),
            ("Promedio (días)", m_emp["antiguedad_prom_dias"]),
        ]),
        ("DOCUMENTOS", [
            ("Total generados", m_doc.get("total", 0)),
            ("Enviados por correo", m_doc.get("enviados", 0)),
        ]),
    ]

    for titulo, items in secciones:
        ws.cell(row=fila, column=1, value=titulo).font = header_font
        ws.cell(row=fila, column=1).fill = header_fill
        ws.cell(row=fila, column=2).fill = header_fill
        fila += 1
        for concepto, valor in items:
            ws.cell(row=fila, column=1, value=concepto)
            ws.cell(row=fila, column=2, value=valor)
            fila += 1
        fila += 1  # espacio

    # Ajustar anchos
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25

    # Guardar en memoria
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    st.download_button(
        "⬇️ Descargar Reporte.xlsx",
        buffer,
        file_name=f"Reporte_GestorRH_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
