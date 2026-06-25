"""Genera la plantilla Base_Empleados.xlsx con todos los campos necesarios."""
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

COLUMNAS = ["Nombre","Documento","Cargo","Salario","Fecha ingreso",
            "Fecha retiro","Tipo contrato","Correo","Cuenta bancaria",
            "Ingreso promedio variable"]
EJEMPLO = [{
    "Nombre":"Juan Pérez","Documento":"1020304050",
    "Cargo":"Auxiliar Administrativo","Salario":1800000,
    "Fecha ingreso":"01/02/2024","Fecha retiro":"",
    "Tipo contrato":"Indefinido","Correo":"empleado@email.com",
    "Cuenta bancaria":"Bco Nacional Cta Ahorros #000111222",
    "Ingreso promedio variable":0,
}]

def crear_plantilla(ruta="plantillas/Base_Empleados.xlsx"):
    df = pd.DataFrame(EJEMPLO, columns=COLUMNAS)
    with pd.ExcelWriter(ruta, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Empleados")
        ws = w.sheets["Empleados"]
        fill = PatternFill(start_color="1B3F6E", end_color="1B3F6E", fill_type="solid")
        font_h = Font(color="FFFFFF", bold=True)
        for i, _ in enumerate(COLUMNAS, 1):
            c = ws.cell(row=1, column=i)
            c.fill = fill; c.font = font_h
            c.alignment = Alignment(horizontal="center")
        anchos = [22,16,24,14,16,16,18,28,32,22]
        for i,a in enumerate(anchos,1):
            ws.column_dimensions[get_column_letter(i)].width = a
        notas = pd.DataFrame({"Instrucciones":[
            "Llena una fila por cada empleado.",
            "Nombre, Documento, Cargo, Salario y Fecha ingreso son OBLIGATORIOS.",
            "Fechas en formato dd/mm/aaaa.",
            "Deja Fecha retiro vacía si el empleado sigue activo.",
            "Tipo contrato: Fijo, Indefinido, Obra o labor, Prestación de servicios.",
            "Salario: número sin puntos ni signos ($).",
            "Cuenta bancaria: para liquidaciones (ej. Bancolombia Ahorros #123456).",
            "Ingreso promedio variable: 0 si no aplica, o el valor promedio mensual.",
        ]})
        notas.to_excel(w, index=False, sheet_name="Instrucciones")
        w.sheets["Instrucciones"].column_dimensions["A"].width = 70
    print(f"Plantilla creada: {ruta}")

if __name__ == "__main__":
    crear_plantilla()
