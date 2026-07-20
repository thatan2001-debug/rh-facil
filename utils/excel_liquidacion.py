"""
Generación de liquidación en formato Excel.

Modo 1 — Formato propio Gestor RH IA:
  Genera un Excel profesional con el cálculo completo.

Modo 2 — Plantilla del usuario:
  El usuario sube su Excel con fórmulas. El sistema detecta las celdas
  de entrada (salario, fechas, días) y las llena automáticamente.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime


# ── Colores por diseño (heredados de plantillas_disenio) ─────────────────────
COLORES_DISENIO = {
    1: {"primario": "1B3F6E", "secundario": "2D6BE4", "acento": "E8F0FD"},
    2: {"primario": "1F2937", "secundario": "B45309", "acento": "FEF3C7"},
    3: {"primario": "064E3B", "secundario": "059669", "acento": "ECFDF5"},
    4: {"primario": "111827", "secundario": "DC2626", "acento": "FEF2F2"},
    5: {"primario": "4C1D95", "secundario": "7C3AED", "acento": "F5F3FF"},
}

COP_FORMAT = '#,##0'


def _fmt_cop(valor):
    try:
        return float(valor)
    except (ValueError, TypeError):
        return 0.0


def _borde_fino():
    lado = Side(style='thin', color="DDDDDD")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _borde_medio():
    lado = Side(style='medium', color="AAAAAA")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def generar_excel_liquidacion_propio(resultado: dict, datos_empresa: dict,
                                      ruta_salida: str, disenio: int = 1):
    """
    Genera un Excel de liquidación con formato profesional propio de Gestor RH IA.
    """
    colores = COLORES_DISENIO.get(disenio, COLORES_DISENIO[1])
    c_prim  = colores["primario"]
    c_sec   = colores["secundario"]
    c_ac    = colores["acento"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liquidación"
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18

    # ── Encabezado empresa ───────────────────────────────────────────────────
    ws.merge_cells("A1:B1")
    ws["A1"] = datos_empresa.get("nombre", "")
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=c_prim)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["C1"] = f"NIT: {datos_empresa.get('nit','')}"
    ws["C1"].font = Font(bold=True, size=10, color="FFFFFF")
    ws["C1"].fill = PatternFill("solid", fgColor=c_prim)
    ws["C1"].alignment = Alignment(horizontal="right", vertical="center")

    # Título
    ws.merge_cells("A2:C2")
    ws["A2"] = "LIQUIDACIÓN DE PRESTACIONES SOCIALES"
    ws["A2"].font = Font(bold=True, size=12, color=c_prim)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    # ── Datos del empleado ───────────────────────────────────────────────────
    fila = 3
    datos_emp = [
        ("Apellidos y Nombre", resultado.get("Nombre", "")),
        ("Cédula No.", resultado.get("Documento", "")),
        ("Cuenta Bancaria", resultado.get("Cuenta bancaria", "")),
        ("Causa Liquidación", resultado.get("Motivo retiro", "Retiro Voluntario")),
        ("Fecha de Retiro", resultado.get("Fecha corte", "")),
        ("Fecha de Ingreso", resultado.get("Fecha ingreso", "")),
        ("Total días a liquidar", resultado.get("Dias totales (base 360)", 0)),
    ]
    for etiqueta, valor in datos_emp:
        ws[f"A{fila}"] = etiqueta
        ws[f"A{fila}"].font = Font(bold=True, size=9, color=c_prim)
        ws[f"A{fila}"].fill = PatternFill("solid", fgColor=c_ac)
        ws[f"A{fila}"].alignment = Alignment(vertical="center")
        ws[f"A{fila}"].border = _borde_fino()
        ws.merge_cells(f"B{fila}:C{fila}")
        ws[f"B{fila}"] = valor
        ws[f"B{fila}"].font = Font(size=9)
        ws[f"B{fila}"].alignment = Alignment(vertical="center")
        ws[f"B{fila}"].border = _borde_fino()
        ws.row_dimensions[fila].height = 18
        fila += 1

    fila += 1  # espacio

    # ── Sueldo mensual ───────────────────────────────────────────────────────
    salario = _fmt_cop(resultado.get("Salario base", 0))
    from utils.calcular_liquidacion import AUXILIO_TRANSPORTE_2026
    auxilio_aplica = resultado.get("Auxilio transporte incluido", "No") == "Sí"
    auxilio_val = AUXILIO_TRANSPORTE_2026 if auxilio_aplica else 0

    for etiqueta, valor, negrita in [
        ("SUELDO MENSUAL", salario, True),
        ("Auxilio de transporte", auxilio_val if auxilio_aplica else "", False),
        ("Base prestacional", salario + auxilio_val, True),
    ]:
        ws[f"A{fila}"] = etiqueta
        ws[f"A{fila}"].font = Font(bold=negrita, size=9)
        ws[f"A{fila}"].border = _borde_fino()
        ws.merge_cells(f"B{fila}:B{fila}")
        ws[f"C{fila}"] = valor if isinstance(valor, (int, float)) else valor
        if isinstance(valor, (int, float)) and valor:
            ws[f"C{fila}"].number_format = COP_FORMAT
        ws[f"C{fila}"].font = Font(bold=negrita, size=9)
        ws[f"C{fila}"].alignment = Alignment(horizontal="right")
        ws[f"C{fila}"].border = _borde_fino()
        if negrita:
            ws[f"A{fila}"].fill = PatternFill("solid", fgColor=c_ac)
            ws[f"C{fila}"].fill = PatternFill("solid", fgColor=c_ac)
        ws.row_dimensions[fila].height = 18
        fila += 1

    fila += 1

    # ── Encabezado tabla prestaciones ────────────────────────────────────────
    for col, texto in [("A", "Concepto"), ("B", "Días"), ("C", "Valor")]:
        ws[f"{col}{fila}"] = texto
        ws[f"{col}{fila}"].font = Font(bold=True, size=9, color="FFFFFF")
        ws[f"{col}{fila}"].fill = PatternFill("solid", fgColor=c_prim)
        ws[f"{col}{fila}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"{col}{fila}"].border = _borde_fino()
    ws.row_dimensions[fila].height = 20
    fila += 1

    # ── Conceptos de prestaciones ────────────────────────────────────────────
    dias_total = resultado.get("Dias totales (base 360)", 0)
    dias_semi  = resultado.get("Dias semestre actual (prima)", 0)
    conceptos  = [
        ("Cesantías (8.33%)", dias_total, resultado.get("Cesantias (Art. 249 CST)", 0)),
        ("Intereses sobre Cesantías (12% anual)", dias_total, resultado.get("Intereses cesantias 12% (Ley 52/75)", 0)),
        ("Vacaciones (4.17%)", dias_total, resultado.get("Vacaciones (Art. 186 CST)", 0)),
        (f"Prima de Servicios (8.33% — {dias_semi} días)", dias_semi, resultado.get("Prima semestral (Art. 306 CST)", 0)),
        ("Salario pendiente del período", "—", resultado.get("Salario pendiente (estimado)", 0)),
    ]
    indem = resultado.get("Indemnizacion (Art. 64 CST)", 0)
    if indem > 0:
        conceptos.append(("Indemnización (Art. 64 CST)", "—", indem))

    for concepto, dias, valor in conceptos:
        ws[f"A{fila}"] = concepto
        ws[f"A{fila}"].font = Font(size=9)
        ws[f"A{fila}"].border = _borde_fino()
        ws[f"B{fila}"] = dias
        ws[f"B{fila}"].font = Font(size=9)
        ws[f"B{fila}"].alignment = Alignment(horizontal="center")
        ws[f"B{fila}"].border = _borde_fino()
        ws[f"C{fila}"] = _fmt_cop(valor)
        ws[f"C{fila}"].number_format = COP_FORMAT
        ws[f"C{fila}"].font = Font(size=9)
        ws[f"C{fila}"].alignment = Alignment(horizontal="right")
        ws[f"C{fila}"].border = _borde_fino()
        ws.row_dimensions[fila].height = 18
        fila += 1

    # Total liquidación
    total_liq = _fmt_cop(resultado.get("TOTAL LIQUIDACION ESTIMADA", 0))
    ws[f"A{fila}"] = "TOTAL LIQUIDACIÓN"
    ws[f"A{fila}"].font = Font(bold=True, size=9, color=c_prim)
    ws[f"A{fila}"].fill = PatternFill("solid", fgColor=c_ac)
    ws[f"A{fila}"].border = _borde_medio()
    ws[f"B{fila}"] = ""
    ws[f"B{fila}"].border = _borde_medio()
    ws[f"C{fila}"] = total_liq
    ws[f"C{fila}"].number_format = COP_FORMAT
    ws[f"C{fila}"].font = Font(bold=True, size=9, color=c_prim)
    ws[f"C{fila}"].fill = PatternFill("solid", fgColor=c_ac)
    ws[f"C{fila}"].alignment = Alignment(horizontal="right")
    ws[f"C{fila}"].border = _borde_medio()
    ws.row_dimensions[fila].height = 20
    fila += 2

    # ── Descuentos seguridad social ──────────────────────────────────────────
    salario_b = _fmt_cop(resultado.get("Salario base", 0))
    eps = round(salario_b * 0.04, 0)
    pen = round(salario_b * 0.04, 0)
    total_desc = eps + pen
    total_neto = round(total_liq - total_desc, 0)

    ws[f"A{fila}"] = "Menos:"
    ws[f"A{fila}"].font = Font(bold=True, size=9)
    ws.row_dimensions[fila].height = 18; fila += 1

    for concepto, porcentaje, valor in [
        ("EPS", "4%", eps), ("Pensión", "4%", pen),
    ]:
        ws[f"A{fila}"] = concepto
        ws[f"A{fila}"].font = Font(size=9)
        ws[f"A{fila}"].border = _borde_fino()
        ws[f"B{fila}"] = porcentaje
        ws[f"B{fila}"].font = Font(size=9)
        ws[f"B{fila}"].alignment = Alignment(horizontal="center")
        ws[f"B{fila}"].border = _borde_fino()
        ws[f"C{fila}"] = valor
        ws[f"C{fila}"].number_format = COP_FORMAT
        ws[f"C{fila}"].font = Font(size=9)
        ws[f"C{fila}"].alignment = Alignment(horizontal="right")
        ws[f"C{fila}"].border = _borde_fino()
        ws.row_dimensions[fila].height = 18; fila += 1

    # Total descuentos
    ws[f"A{fila}"] = "TOTAL DESCUENTOS SEGURIDAD SOCIAL"
    ws[f"A{fila}"].font = Font(bold=True, size=9)
    ws[f"A{fila}"].border = _borde_medio()
    ws[f"C{fila}"] = total_desc
    ws[f"C{fila}"].number_format = COP_FORMAT
    ws[f"C{fila}"].font = Font(bold=True, size=9)
    ws[f"C{fila}"].alignment = Alignment(horizontal="right")
    ws[f"C{fila}"].border = _borde_medio()
    ws.row_dimensions[fila].height = 20; fila += 1

    # Neto
    ws[f"A{fila}"] = "LIQUIDACIÓN NETA A PAGAR"
    ws[f"A{fila}"].font = Font(bold=True, size=10, color="FFFFFF")
    ws[f"A{fila}"].fill = PatternFill("solid", fgColor=c_sec)
    ws[f"A{fila}"].border = _borde_medio()
    ws[f"B{fila}"] = ""
    ws[f"B{fila}"].fill = PatternFill("solid", fgColor=c_sec)
    ws[f"C{fila}"] = total_neto
    ws[f"C{fila}"].number_format = COP_FORMAT
    ws[f"C{fila}"].font = Font(bold=True, size=10, color="FFFFFF")
    ws[f"C{fila}"].fill = PatternFill("solid", fgColor=c_sec)
    ws[f"C{fila}"].alignment = Alignment(horizontal="right")
    ws[f"C{fila}"].border = _borde_medio()
    ws.row_dimensions[fila].height = 22; fila += 2

    # ── Firmas ───────────────────────────────────────────────────────────────
    ws[f"A{fila}"] = "_" * 35
    ws[f"C{fila}"] = "_" * 35
    ws.row_dimensions[fila].height = 18; fila += 1

    ws[f"A{fila}"] = datos_empresa.get("representante", "Representante Legal")
    ws[f"A{fila}"].font = Font(bold=True, size=9)
    ws[f"C{fila}"] = resultado.get("Nombre", "Empleado")
    ws[f"C{fila}"].font = Font(bold=True, size=9)
    ws.row_dimensions[fila].height = 16; fila += 1

    ws[f"A{fila}"] = f"Representante Legal — {datos_empresa.get('nombre','')}"
    ws[f"A{fila}"].font = Font(size=8, color="888888")
    ws[f"C{fila}"] = f"C.C.: {resultado.get('Documento','')}"
    ws[f"C{fila}"].font = Font(size=8, color="888888")
    ws.row_dimensions[fila].height = 15; fila += 2

    # Aviso legal
    ws.merge_cells(f"A{fila}:C{fila}")
    ws[f"A{fila}"] = (
        "AVISO: Liquidación ESTIMADA (base 360 días). "
        "Validar con contador antes del pago."
    )
    ws[f"A{fila}"].font = Font(italic=True, size=8, color="888888")
    ws[f"A{fila}"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[fila].height = 20

    # Proteger hoja de cálculos (solo lectura para el usuario)
    ws.protection.sheet = True
    ws.protection.password = "rhfacil"

    wb.save(ruta_salida)


def llenar_plantilla_usuario(ruta_plantilla: str, resultado: dict,
                              ruta_salida: str) -> tuple[bool, str]:
    """
    Intenta llenar la plantilla Excel del usuario con los datos calculados.
    
    Busca celdas que contengan etiquetas reconocibles y las llena con los valores.
    Retorna (éxito, mensaje).
    """
    try:
        wb = openpyxl.load_workbook(ruta_plantilla)
        ws = wb.active

        from utils.calcular_liquidacion import AUXILIO_TRANSPORTE_2026
        auxilio_aplica = resultado.get("Auxilio transporte incluido", "No") == "Sí"

        # Mapa de palabras clave → valores a insertar
        mapa_valores = {
            # Salario
            "salario": float(resultado.get("Salario base", 0)),
            "sueldo": float(resultado.get("Salario base", 0)),
            # Datos empleado
            "nombre": resultado.get("Nombre", ""),
            "cedula": resultado.get("Documento", ""),
            "cédula": resultado.get("Documento", ""),
            "cargo": resultado.get("Cargo", ""),
            "cuenta": resultado.get("Cuenta bancaria", ""),
            # Fechas
            "ingreso": resultado.get("Fecha ingreso", ""),
            "retiro": resultado.get("Fecha corte", ""),
            # Días
            "dias": resultado.get("Dias totales (base 360)", 0),
            "días": resultado.get("Dias totales (base 360)", 0),
            # Prestaciones calculadas
            "cesantia": float(resultado.get("Cesantias (Art. 249 CST)", 0)),
            "cesantía": float(resultado.get("Cesantias (Art. 249 CST)", 0)),
            "interes": float(resultado.get("Intereses cesantias 12% (Ley 52/75)", 0)),
            "interés": float(resultado.get("Intereses cesantias 12% (Ley 52/75)", 0)),
            "vacacion": float(resultado.get("Vacaciones (Art. 186 CST)", 0)),
            "vacación": float(resultado.get("Vacaciones (Art. 186 CST)", 0)),
            "prima": float(resultado.get("Prima semestral (Art. 306 CST)", 0)),
            "auxilio": AUXILIO_TRANSPORTE_2026 if auxilio_aplica else 0,
        }

        celdas_llenadas = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val_lower = cell.value.lower().strip()
                    for clave, valor in mapa_valores.items():
                        if clave in val_lower:
                            # La celda a la derecha recibe el valor
                            celda_valor = ws.cell(
                                row=cell.row,
                                column=cell.column + 1
                            )
                            if celda_valor.value is None or celda_valor.value == "":
                                celda_valor.value = valor
                                if isinstance(valor, float):
                                    celda_valor.number_format = '#,##0'
                                celdas_llenadas += 1
                            break

        wb.save(ruta_salida)
        return True, f"Plantilla llenada: {celdas_llenadas} celda(s) completada(s)."
    except Exception as e:
        return False, f"Error llenando plantilla: {e}"
