"""
Generación de liquidación en formato Excel.
Tres modos:
1. Formato simple: tabla limpia con los valores
2. Con fórmulas: celdas de entrada + fórmulas Excel reales
3. Plantilla propia: el usuario sube su Excel y el sistema rellena celdas clave
"""

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from pathlib import Path

AZUL    = "1B3F6E"
AZUL_CL = "DBEAFE"
GRIS    = "F2F4F7"
VERDE   = "065F46"
VERDE_B = "D1FAE5"
ROJO_B  = "FEE2E2"

def _borde_fino():
    lado = Side(style="thin", color="DDDDDD")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def exportar_excel_simple(resultados: list[dict], ruta_salida: str):
    """Genera un Excel limpio con una fila por empleado."""
    columnas = [
        "Nombre", "Documento", "Cargo", "Fecha ingreso", "Fecha corte",
        "Dias totales (base 360)", "Salario base",
        "Cesantias (Art. 249 CST)", "Intereses cesantias 12% (Ley 52/75)",
        "Vacaciones (Art. 186 CST)", "Prima semestral (Art. 306 CST)",
        "Salario pendiente (estimado)", "Indemnizacion (Art. 64 CST)",
        "TOTAL LIQUIDACION ESTIMADA",
        "Descuento EPS (4%)", "Descuento Pension (4%)",
        "TOTAL NETO A PAGAR",
    ]
    filas = []
    for r in resultados:
        salario = float(r.get("Salario base", 0))
        total   = float(r.get("TOTAL LIQUIDACION ESTIMADA", 0))
        eps     = round(salario * 0.04, 2)
        pension = round(salario * 0.04, 2)
        neto    = round(total - eps - pension, 2)
        filas.append({
            "Nombre":                              r.get("Nombre",""),
            "Documento":                           r.get("Documento",""),
            "Cargo":                               r.get("Cargo",""),
            "Fecha ingreso":                       r.get("Fecha ingreso",""),
            "Fecha corte":                         r.get("Fecha corte",""),
            "Dias totales (base 360)":             r.get("Dias totales (base 360)",0),
            "Salario base":                        salario,
            "Cesantias (Art. 249 CST)":            r.get("Cesantias (Art. 249 CST)",0),
            "Intereses cesantias 12% (Ley 52/75)": r.get("Intereses cesantias 12% (Ley 52/75)",0),
            "Vacaciones (Art. 186 CST)":           r.get("Vacaciones (Art. 186 CST)",0),
            "Prima semestral (Art. 306 CST)":      r.get("Prima semestral (Art. 306 CST)",0),
            "Salario pendiente (estimado)":        r.get("Salario pendiente (estimado)",0),
            "Indemnizacion (Art. 64 CST)":         r.get("Indemnizacion (Art. 64 CST)",0),
            "TOTAL LIQUIDACION ESTIMADA":          total,
            "Descuento EPS (4%)":                  eps,
            "Descuento Pension (4%)":              pension,
            "TOTAL NETO A PAGAR":                  neto,
        })

    df = pd.DataFrame(filas, columns=columnas)
    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Liquidaciones")
        ws = w.sheets["Liquidaciones"]

        # Encabezado
        fill_h = PatternFill("solid", fgColor=AZUL)
        font_h = Font(color="FFFFFF", bold=True, size=9)
        for col_idx in range(1, len(columnas)+1):
            c = ws.cell(row=1, column=col_idx)
            c.fill = fill_h; c.font = font_h
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        # Filas de datos
        fill_alt  = PatternFill("solid", fgColor=GRIS)
        fill_tot  = PatternFill("solid", fgColor=VERDE_B)
        font_tot  = Font(bold=True, color=VERDE)
        fmt_peso  = '#,##0'
        col_dinero = list(range(7, len(columnas)+1))  # columnas con valores $

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            es_par = (row_idx % 2 == 0)
            for col_idx, cell in enumerate(row, start=1):
                cell.border    = _borde_fino()
                cell.font      = Font(size=9)
                cell.alignment = Alignment(horizontal="center")
                if es_par:
                    cell.fill = fill_alt
                if col_idx in col_dinero and cell.value is not None:
                    cell.number_format = fmt_peso

        # Resaltar columna TOTAL NETO
        col_neto = len(columnas)
        for row_idx in range(2, len(filas)+2):
            c = ws.cell(row=row_idx, column=col_neto)
            c.fill = fill_tot; c.font = font_tot

        # Fila de totales
        fila_tot = len(filas) + 2
        ws.cell(fila_tot, 1, "TOTALES").font = Font(bold=True, size=9)
        for col_idx in col_dinero:
            letra = get_column_letter(col_idx)
            c = ws.cell(fila_tot, col_idx)
            c.value  = f"=SUM({letra}2:{letra}{fila_tot-1})"
            c.fill   = PatternFill("solid", fgColor=AZUL_CL)
            c.font   = Font(bold=True, size=9)
            c.number_format = fmt_peso
            c.border = _borde_fino()

        # Hoja de aviso legal
        ws_av = w.book.create_sheet("Aviso Legal")
        avisos = [
            "AVISO IMPORTANTE — Gestor RH IA",
            "",
            "Las liquidaciones de esta hoja son ESTIMACIONES DE REFERENCIA.",
            "Fórmulas usadas (base año comercial 360 días):",
            "  • Cesantías = Salario base × días / 360",
            "  • Intereses cesantías = Cesantías × 12% × días / 360",
            "  • Vacaciones = Salario × días / 720",
            "  • Prima semestral = Salario base × días semestre / 360",
            "",
            "NO incluye: salario integral, mora en cesantías, incapacidades,",
            "embargos, horas extras, sanciones ni casos especiales.",
            "",
            "Valide siempre con su contador o abogado laboral antes de realizar pagos.",
            "Salario mínimo 2026: $1.750.905 | Auxilio transporte: $249.095",
        ]
        for i, texto in enumerate(avisos, 1):
            c = ws_av.cell(row=i, column=1, value=texto)
            if i == 1:
                c.font = Font(bold=True, size=12, color=AZUL)
            ws_av.column_dimensions["A"].width = 70


def exportar_excel_con_formulas(resultados: list[dict], ruta_salida: str):
    """
    Genera Excel con celdas de entrada y fórmulas reales de Excel.
    Cada empleado tiene su propia sección con fórmulas transparentes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Liquidaciones"

    fill_h   = PatternFill("solid", fgColor=AZUL)
    fill_sec = PatternFill("solid", fgColor=AZUL_CL)
    fill_res = PatternFill("solid", fgColor=VERDE_B)
    fill_des = PatternFill("solid", fgColor=ROJO_B)
    font_h   = Font(color="FFFFFF", bold=True, size=10)
    font_b   = Font(bold=True, size=10)
    fmt_peso = '#,##0'

    # Título
    ws.merge_cells("A1:H1")
    ws["A1"] = "LIQUIDACIÓN DE PRESTACIONES SOCIALES — Gestor RH IA"
    ws["A1"].font  = Font(bold=True, size=14, color=AZUL)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    fila = 3
    for r in resultados:
        salario   = float(r.get("Salario base", 0))
        dias      = r.get("Dias totales (base 360)", 0)
        dias_semi = r.get("Dias semestre actual (prima)", 0)

        # Encabezado empleado
        ws.merge_cells(f"A{fila}:H{fila}")
        ws[f"A{fila}"] = f"Empleado: {r.get('Nombre','')} — C.C. {r.get('Documento','')}"
        ws[f"A{fila}"].fill = fill_h; ws[f"A{fila}"].font = font_h
        ws[f"A{fila}"].alignment = Alignment(horizontal="left")
        ws.row_dimensions[fila].height = 20
        fila += 1

        # Datos de entrada (celdas amarillas = editables)
        entradas = [
            ("Salario base mensual ($)", salario, "D"),
            ("Días trabajados (base 360)", dias, "E"),
            ("Días semestre actual (prima)", dias_semi, "F"),
        ]
        fill_ent = PatternFill("solid", fgColor="FEF9C3")
        for label, valor, _ in entradas:
            ws[f"B{fila}"] = label
            ws[f"B{fila}"].font = Font(size=9, bold=True)
            ws[f"C{fila}"] = valor
            ws[f"C{fila}"].fill = fill_ent
            ws[f"C{fila}"].number_format = fmt_peso
            ws[f"C{fila}"].font = Font(size=9)
            ws[f"C{fila}"].border = _borde_fino()
            fila += 1

        # Referencia a celdas de entrada
        r_sal  = fila - 3  # fila del salario
        r_dias = fila - 2  # fila de días
        r_semi = fila - 1  # fila días semestre

        # Conceptos con fórmulas
        conceptos = [
            ("Cesantías (8.33% — Art. 249 CST)",
             f"=C{r_sal}*C{r_dias}/360"),
            ("Intereses Cesantías (12% — Ley 52/75)",
             f"=D{fila}*0.12*C{r_dias}/360"),
            ("Vacaciones (4.17% — Art. 186 CST)",
             f"=C{r_sal}*C{r_dias}/720"),
            ("Prima Semestral (8.33% — Art. 306 CST)",
             f"=C{r_sal}*C{r_semi}/360"),
            ("Salario pendiente del período",
             float(r.get("Salario pendiente (estimado)", 0))),
        ]
        if float(r.get("Indemnizacion (Art. 64 CST)", 0)) > 0:
            conceptos.append(("Indemnización (Art. 64 CST)",
                float(r.get("Indemnizacion (Art. 64 CST)", 0))))

        filas_conceptos = []
        for label, formula in conceptos:
            ws[f"B{fila}"] = label
            ws[f"B{fila}"].font = Font(size=9)
            ws[f"B{fila}"].border = _borde_fino()
            ws[f"D{fila}"] = formula
            ws[f"D{fila}"].number_format = fmt_peso
            ws[f"D{fila}"].border = _borde_fino()
            ws[f"D{fila}"].font = Font(size=9)
            filas_conceptos.append(fila)
            fila += 1

        # Subtotal
        rango = f"D{filas_conceptos[0]}:D{filas_conceptos[-1]}"
        ws[f"B{fila}"] = "TOTAL LIQUIDACIÓN"
        ws[f"B{fila}"].font = font_b; ws[f"B{fila}"].fill = fill_res
        ws[f"D{fila}"] = f"=SUM({rango})"
        ws[f"D{fila}"].number_format = fmt_peso
        ws[f"D{fila}"].font = font_b; ws[f"D{fila}"].fill = fill_res
        ws[f"D{fila}"].border = _borde_fino()
        fila_tot = fila; fila += 1

        # Descuentos SS
        ws[f"B{fila}"] = "Descuento EPS (4%)"
        ws[f"B{fila}"].font = Font(size=9)
        ws[f"D{fila}"] = f"=C{r_sal}*0.04"
        ws[f"D{fila}"].number_format = fmt_peso
        ws[f"D{fila}"].fill = fill_des; ws[f"D{fila}"].border = _borde_fino()
        fila_eps = fila; fila += 1

        ws[f"B{fila}"] = "Descuento Pensión (4%)"
        ws[f"B{fila}"].font = Font(size=9)
        ws[f"D{fila}"] = f"=C{r_sal}*0.04"
        ws[f"D{fila}"].number_format = fmt_peso
        ws[f"D{fila}"].fill = fill_des; ws[f"D{fila}"].border = _borde_fino()
        fila_pen = fila; fila += 1

        ws[f"B{fila}"] = "TOTAL NETO A PAGAR"
        ws[f"B{fila}"].font = Font(bold=True, size=10, color=VERDE)
        ws[f"D{fila}"] = f"=D{fila_tot}-D{fila_eps}-D{fila_pen}"
        ws[f"D{fila}"].number_format = fmt_peso
        ws[f"D{fila}"].font = Font(bold=True, size=10, color=VERDE)
        ws[f"D{fila}"].fill = fill_res; ws[f"D{fila}"].border = _borde_fino()
        fila += 2  # Separador entre empleados

    # Anchos
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    ws_av = wb.create_sheet("Instrucciones")
    ws_av["A1"] = "Las celdas en AMARILLO son editables. Los demás valores se calculan automáticamente con fórmulas."
    ws_av["A1"].font = Font(bold=True, size=11)
    ws_av["A2"] = "Puedes ajustar el salario, días trabajados o días de semestre para cada empleado."
    ws_av["A3"] = "AVISO: Esta es una estimación. Valide con su contador antes de realizar pagos."
    ws_av.column_dimensions["A"].width = 80

    wb.save(ruta_salida)
